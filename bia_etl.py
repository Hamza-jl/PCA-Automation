#!/usr/bin/env python3
"""
BIA ETL Pipeline
================
Extracts structured data from BIA fiche (.docx) files and fills the
Synthèse BIA (.xlsx) template.

WHY NO LLM:
  - All source data lives in Word tables, not free text.
  - Mapping fiche-table → synthesis-sheet is deterministic.
  - Validations are finite lookup dictionaries.
  - LLMs would add hallucination risk, cost, and non-reproducibility
    for a task that is 100% structural/formulaic.

WHAT IS "AI" HERE:
  - rapidfuzz: fuzzy string matching for department name lookup.
    The synthesis template has typos ("ingenieurue") vs the fiche's
    correct spelling. A simple Levenshtein ratio resolves this safely.

USAGE:
  # Single fiche:
  python bia_etl.py --fiche path/to/fiche.docx --synthese path/to/synthese.xlsx

  # Folder of fiches:
  python bia_etl.py --fiches-dir path/to/fiches/ --synthese path/to/synthese.xlsx

  # Dry-run (extract + print, no write):
  python bia_etl.py --fiche path/to/fiche.docx --synthese path/to/synthese.xlsx --dry-run
"""

import argparse
import copy
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process


# ─────────────────────────────────────────────────────────────────────────────
# DATA CLASSES  (the "schema" of a BIA fiche)
# ─────────────────────────────────────────────────────────────────────────────
# These are plain dataclasses — no ORM, no DB, no serialisation overhead.
# They represent exactly what can be extracted from a fiche word-for-word.

@dataclass
class Activity:
    name: str
    resources: str
    critical_period: str
    criticality: str

@dataclass
class ImpactRow:
    """One activity's impact scores across all dimensions and time scenarios."""
    activity_name: str
    # scenario A = interruption < 1 day  →  maps to 1H, 4H, 1J columns
    # scenario B = interruption ≥ 5 days →  maps to 2-3J column
    # WHY this mapping: the fiche captures 2 scenarios, synthesis needs 4.
    # We propagate A across the 3 short-horizon columns and use B for the
    # 2-3J column. This is conservative and flagged in comments for human review.
    im_a: str; im_b: str   # Image de marque
    di_a: str; di_b: str   # Désorganisation interne
    jr_a: str; jr_b: str   # Juridique / Réglementaire
    fin_a: str; fin_b: str # Financier
    dmia_expressed: str = ""

@dataclass
class Exchange:
    correspondent: str
    ie_type: str    # "I" = interne, "E" = externe
    info_type: str
    tr_type: str    # "T" = transmis, "R" = reçu
    si_resources: str

@dataclass
class RampUpRow:
    """One row of the montée en charge table (Effectif, Positions, Télétravail…)."""
    label: str
    nominal: str
    h0: str = ""; h1: str = ""; h2: str = ""; h4: str = ""
    j1: str = ""; j2: str = ""; j3: str = ""; j4: str = ""; j5: str = ""
    j10: str = ""; j15: str = ""; j30: str = ""
    comments: str = ""

@dataclass
class KeyPerson:
    function: str
    last_name: str
    first_name: str
    position: str = ""
    seniority: str = ""
    replacements: str = ""

@dataclass
class ITApplication:
    name: str
    criticality: str
    dmia: str
    pmdt: str
    workaround: str = ""
    comments: str = ""

@dataclass
class OtherEquipment:
    designation: str
    h0: str = ""; h1: str = ""; h2: str = ""; h4: str = ""
    j1: str = ""; j2: str = ""; j3: str = ""; j5: str = ""
    j10: str = ""; j15: str = ""
    comments: str = ""

@dataclass
class CriticalDoc:
    name: str
    storage_type: str
    duplication: str
    duplication_method: str = ""

@dataclass
class BIAFiche:
    """All structured data extracted from one BIA fiche word document."""
    entity_name: str
    division: str = ""
    unite: str = ""
    department: str = ""
    activities: list[Activity] = field(default_factory=list)
    impact_rows: list[ImpactRow] = field(default_factory=list)
    exchanges: list[Exchange] = field(default_factory=list)
    ramp_up: list[RampUpRow] = field(default_factory=list)
    key_people: list[KeyPerson] = field(default_factory=list)
    it_applications: list[ITApplication] = field(default_factory=list)
    other_equipment: list[OtherEquipment] = field(default_factory=list)
    critical_docs: list[CriticalDoc] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# IMPACT SCORE CALCULATION
# ─────────────────────────────────────────────────────────────────────────────
# Weights are hardcoded from the impact scale table in the fiche:
#   Image de marque = 4, Désorganisation interne = 1, Réglementaire = 2, Financier = 3
# Score = sum(impact_level × weight) for all 4 dimensions.
# The Echelle d'impact sheet then maps the score to Faible/Significatif/Majeur/Catastrophique.

IMPACT_WEIGHTS = {"im": 4, "di": 1, "jr": 2, "fin": 3}

def _safe_int(v: str) -> Optional[int]:
    """Convert impact cell value to int, returning None for NA/empty/non-numeric."""
    if not v or v.strip().upper() in ("NA", "-", ""):
        return None
    try:
        return int(v.strip())
    except ValueError:
        return None

def compute_score(im: str, di: str, jr: str, fin: str) -> str:
    """Compute weighted impact score. Returns empty string if any dimension is missing."""
    vals = [_safe_int(im), _safe_int(di), _safe_int(jr), _safe_int(fin)]
    if None in vals:
        return ""
    return str(vals[0]*4 + vals[1]*1 + vals[2]*2 + vals[3]*3)


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACT — DOCX PARSER
# ─────────────────────────────────────────────────────────────────────────────
# WHY semantic detection instead of fixed indices:
#   Different fiches have different numbers of activities → different numbers
#   of impact matrix tables → table indices shift. Using column headers as
#   fingerprints makes the parser robust to any fiche in this family.

def _cell(table, row_i: int, col_i: int) -> str:
    """Safe cell accessor — returns empty string on index error."""
    try:
        return table.rows[row_i].cells[col_i].text.strip()
    except IndexError:
        return ""

def _row_texts(row) -> list[str]:
    return [c.text.strip() for c in row.cells]

def _table_fingerprint(table) -> str:
    """First non-empty cell in row 0 — used as a stable table identifier."""
    for row in table.rows:
        for cell in row.cells:
            text = cell.text.strip()
            if text:
                return text
    return ""

def _is_activity_list_table(table) -> bool:
    if len(table.columns) < 4:
        return False
    headers = _row_texts(table.rows[0])
    return any("Activit" in h for h in headers) and any("Ressource" in h for h in headers)

def _is_impact_matrix_table(table) -> bool:
    """Impact per activity: 3 cols, header row has 'Interruption' in col 1 or 2."""
    if len(table.rows) < 4:
        return False
    headers = _row_texts(table.rows[0])
    return len(headers) >= 3 and "Interruption" in " ".join(headers)

def _is_dmia_table(table) -> bool:
    headers = _row_texts(table.rows[0])
    return any("DMIA" in h for h in headers) and any("Processus" in h or "Process" in h for h in headers)

def _is_exchange_table(table) -> bool:
    headers = _row_texts(table.rows[0])
    return any("Groupes fonctionnels" in h or "Correspondants" in h for h in headers)

def _is_ramp_up_table(table) -> bool:
    if len(table.rows) < 2:
        return False
    headers = _row_texts(table.rows[0])
    return any("Montée en charge" in h for h in headers) and any("Nominal" in h for h in headers)

def _is_key_people_table(table) -> bool:
    headers = _row_texts(table.rows[0])
    return any("Fonction" in h for h in headers) and any("Nom" in h for h in headers) and any("Prénom" in h or "Prenom" in h for h in headers)

def _is_app_table(table) -> bool:
    headers = _row_texts(table.rows[0])
    return any("Application" in h for h in headers) and any("DMIA" in h for h in headers) and any("PMDT" in h for h in headers)

def _is_other_eqt_table(table) -> bool:
    if len(table.rows) < 2:
        return False
    headers = _row_texts(table.rows[0])
    return any("Désignation" in h or "Designation" in h for h in headers) and any("H+" in h or "J+" in h for h in headers)

def _is_doc_table(table) -> bool:
    headers = _row_texts(table.rows[0])
    return any("Documents" in h or "Fichiers" in h for h in headers) and any("Stockage" in h for h in headers)

def _is_identification_table(table) -> bool:
    if len(table.rows) < 2:
        return False
    fp = _table_fingerprint(table)
    return "Entit" in fp or ("Présents" in fp)

def extract(docx_path: str | Path) -> BIAFiche:
    """
    EXTRACT phase: reads a .docx fiche and returns a BIAFiche dataclass.

    Table detection uses column header fingerprints, not positional indices,
    so it works regardless of how many activities (and thus impact tables) the
    fiche contains.
    """
    doc = Document(str(docx_path))
    fiche = BIAFiche(entity_name="")

    # Collect impact matrix tables separately (one per activity, identified together)
    raw_impact_tables = []
    dmia_map: dict[str, str] = {}  # activity_name → dmia_expressed

    for table in doc.tables:
        if _is_identification_table(table) and not fiche.entity_name:
            # Row 0 col 1 = entity name
            fiche.entity_name = _cell(table, 0, 1) or _cell(table, 0, 2)

        elif _is_activity_list_table(table):
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 4 and cells[0]:
                    fiche.activities.append(Activity(
                        name=cells[0],
                        resources=cells[1],
                        critical_period=cells[2],
                        criticality=cells[3],
                    ))

        elif _is_impact_matrix_table(table):
            raw_impact_tables.append(table)

        elif _is_dmia_table(table):
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 2 and cells[0]:
                    dmia_map[cells[0].strip()] = cells[1].strip()

        elif _is_exchange_table(table):
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 5 and cells[0]:
                    fiche.exchanges.append(Exchange(
                        correspondent=cells[0],
                        ie_type=cells[1],
                        info_type=cells[2],
                        tr_type=cells[3],
                        si_resources=cells[4],
                    ))

        elif _is_ramp_up_table(table):
            # Row 0 = header with time horizons
            # Rows 1+ = Effectif, Positions, Télétravail, ...
            headers = _row_texts(table.rows[0])

            def _col_idx(keyword: str) -> int:
                for i, h in enumerate(headers):
                    if keyword in h:
                        return i
                return -1

            for row in table.rows[1:]:
                cells = _row_texts(row)
                if not cells or not cells[0]:
                    continue

                def _get(keyword: str) -> str:
                    idx = _col_idx(keyword)
                    return cells[idx] if 0 <= idx < len(cells) else ""

                fiche.ramp_up.append(RampUpRow(
                    label=cells[0],
                    nominal=_get("Nominal"),
                    h0=_get("H0"),
                    h1=_get("H+1"),
                    h2=_get("H+2"),
                    h4=_get("H+4"),
                    j1=_get("J+1"),
                    j2=_get("J+2"),
                    j3=_get("J+3"),
                    j4=_get("J+4"),
                    j5=_get("J+5"),
                    j10=_get("J+10"),
                    j15=_get("J+15"),
                    j30=_get("J+30"),
                    comments=cells[-1] if len(cells) > 2 else "",
                ))

        elif _is_key_people_table(table):
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 3 and (cells[1] or cells[2]):
                    fiche.key_people.append(KeyPerson(
                        function=cells[0],
                        last_name=cells[1],
                        first_name=cells[2],
                        replacements=cells[3] if len(cells) > 3 else "",
                    ))

        elif _is_app_table(table):
            headers = _row_texts(table.rows[0])
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 4 and cells[0]:
                    fiche.it_applications.append(ITApplication(
                        name=cells[0],
                        criticality=cells[1],
                        dmia=cells[2],
                        pmdt=cells[3],
                        comments=cells[4] if len(cells) > 4 else "",
                    ))

        elif _is_other_eqt_table(table):
            headers = _row_texts(table.rows[0])
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if not cells or not cells[0]:
                    continue

                def _get_eqt(keyword: str) -> str:
                    for i, h in enumerate(headers):
                        if keyword in h:
                            return cells[i] if i < len(cells) else ""
                    return ""

                fiche.other_equipment.append(OtherEquipment(
                    designation=cells[0],
                    h0=_get_eqt("H0"),
                    h2=_get_eqt("H+2"),
                    h4=_get_eqt("H+4"),
                    j1=_get_eqt("J+1"),
                    j2=_get_eqt("J+2"),
                    j3=_get_eqt("J+3"),
                    j5=_get_eqt("J+5"),
                    j10=_get_eqt("J+10"),
                    j15=_get_eqt("J+15"),
                ))

        elif _is_doc_table(table):
            for row in table.rows[1:]:
                cells = _row_texts(row)
                if len(cells) >= 3 and cells[0]:
                    fiche.critical_docs.append(CriticalDoc(
                        name=cells[0],
                        storage_type=cells[1],
                        duplication=cells[2],
                        duplication_method=cells[2] if len(cells) > 3 else cells[2],
                    ))

    # Process impact matrices: each table = one activity's impact scores
    # WHY we do this after the main loop: we need dmia_map to be populated first
    for t in raw_impact_tables:
        activity_name = _cell(t, 0, 0)
        scores: dict[str, dict[str, str]] = {}
        for row_idx, row in enumerate(t.rows[1:], start=1):
            dim_name = _cell(t, row_idx, 0).lower()
            val_a = _cell(t, row_idx, 1)
            val_b = _cell(t, row_idx, 2)
            if "image" in dim_name or "marque" in dim_name:
                scores["im"] = {"a": val_a, "b": val_b}
            elif "d" in dim_name and ("sorg" in dim_name or "org" in dim_name):
                scores["di"] = {"a": val_a, "b": val_b}
            elif "r" in dim_name and ("gl" in dim_name or "juridique" in dim_name):
                scores["jr"] = {"a": val_a, "b": val_b}
            elif "financ" in dim_name:
                scores["fin"] = {"a": val_a, "b": val_b}

        def _s(dim: str, scenario: str) -> str:
            return scores.get(dim, {}).get(scenario, "")

        fiche.impact_rows.append(ImpactRow(
            activity_name=activity_name,
            im_a=_s("im","a"), im_b=_s("im","b"),
            di_a=_s("di","a"), di_b=_s("di","b"),
            jr_a=_s("jr","a"), jr_b=_s("jr","b"),
            fin_a=_s("fin","a"), fin_b=_s("fin","b"),
            dmia_expressed=dmia_map.get(activity_name, ""),
        ))

    # Resolve Division / Unité / Département from entity name
    # Convention in the fiche: entity_name IS the département (or unité if top-level)
    fiche.department = fiche.entity_name

    return fiche


# ─────────────────────────────────────────────────────────────────────────────
# TRANSFORM — SHEET DATA BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
# Each function takes a BIAFiche and returns a list of dicts, where each dict
# maps column_name → value. The LOAD phase writes these into the XLSX.

def transform_activites(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Activité": a.name,
            "Ressources Utilisées": a.resources,
            "Période critique": a.critical_period,
            "Niveau de criticité": a.criticality,
            "Volume": "",
            "Commentaires": "",
        }
        for a in fiche.activities
    ]

def transform_impact_dmia(fiche: BIAFiche) -> list[dict]:
    rows = []
    for ir in fiche.impact_rows:
        # Scenario A (<1J) → propagated to 1H, 4H, 1J (conservative: short outage)
        # Scenario B (≥5J) → used for 2-3J (closest available long-horizon data)
        # This is a deliberate approximation — flag for human review
        score_a = compute_score(ir.im_a, ir.di_a, ir.jr_a, ir.fin_a)
        score_b = compute_score(ir.im_b, ir.di_b, ir.jr_b, ir.fin_b)
        rows.append({
            "Activité": ir.activity_name,
            "IM 1H": ir.im_a, "IM 4H": ir.im_a, "IM 1J": ir.im_a, "IM 2-3J": ir.im_b,
            "DI 1H": ir.di_a, "DI 4H": ir.di_a, "DI 1J": ir.di_a, "DI 2-3J": ir.di_b,
            "JR 1H": ir.jr_a, "JR 4H": ir.jr_a, "JR 1J": ir.jr_a, "JR 2-3J": ir.jr_b,
            "FIN 1H": ir.fin_a, "FIN 4H": ir.fin_a, "FIN 1J": ir.fin_a, "FIN 2-3J": ir.fin_b,
            "Score 1H": score_a, "Score 4H": score_a, "Score 1J": score_a, "Score 2-3J": score_b,
            "DMIA Exprimée": ir.dmia_expressed,
            "DMIA Préconisé": "",
            "Commentaires": "[AUTO] Scores scenario A propagated to 1H/4H/1J columns - verify with client",
        })
    return rows

def transform_applications_it(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Application": a.name,
            "Niveau de criticité": a.criticality,
            "DMIA": a.dmia,
            "PMDT": a.pmdt,
            "Contournement envisageable": a.workaround,
            "Commentaires": a.comments,
        }
        for a in fiche.it_applications
    ]

def transform_echanges_internes(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Groupes fonctionnels / Correspondants": e.correspondent,
            "Types d'informations": e.info_type,
            "Niveau de criticité": "",
            "T / R": e.tr_type,
            "Ressources utilisées": e.si_resources,
            "Commentaires": "",
        }
        for e in fiche.exchanges if e.ie_type.strip().upper() == "I"
    ]

def transform_echanges_externes(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Groupes fonctionnels / Correspondants": e.correspondent,
            "Types d'informations": e.info_type,
            "Typologie": "",
            "T / R": e.tr_type,
            "Ressources utilisées": e.si_resources,
            "Commentaires": "",
        }
        for e in fiche.exchanges if e.ie_type.strip().upper() == "E"
    ]

def transform_montee_en_charge(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Montée en charge exprimée": r.label,
            "Nominal": r.nominal,
            "H0": r.h0, "H+1": r.h1, "H+2": r.h2, "H+4": r.h4,
            "J+1": r.j1, "J+2": r.j2, "J+3": r.j3, "J+4": r.j4,
            "J+5": r.j5, "J+10": r.j10, "J+15": r.j15, "J+30": r.j30,
            "Commentaires": r.comments,
        }
        for r in fiche.ramp_up
    ]

def transform_collaborateurs(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Fonction": p.function,
            "Nom": p.last_name,
            "Prénom": p.first_name,
            "Poste": p.position or p.function,
            "Ancienneté dans le poste": p.seniority,
            "Suppléants possibles": p.replacements,
            "Commentaires": "",
        }
        for p in fiche.key_people
    ]

def transform_autres_eqt(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Désignation": e.designation,
            "H0": e.h0, "H+1": e.h1, "H+2": e.h2, "H+4": e.h4,
            "J+1": e.j1, "J+2": e.j2, "J+3": e.j3, "J+5": e.j5,
            "J+10": e.j10, "J+15": e.j15,
            "Commentaires": e.comments,
        }
        for e in fiche.other_equipment
    ]

def transform_doc_critiques(fiche: BIAFiche) -> list[dict]:
    return [
        {
            "Documents / Fichiers": d.name,
            "Type de stockage\n(Electronique / Papier)": d.storage_type,
            "Duplication \n(O / N)": d.duplication,
            "Modalité de Duplication": d.duplication_method,
            "Commentaires": "",
        }
        for d in fiche.critical_docs
    ]

TRANSFORM_MAP = {
    "Activités":           transform_activites,
    "Impact DMIA":         transform_impact_dmia,
    "Applications IT":     transform_applications_it,
    "Echanges I":          transform_echanges_internes,
    "Echanges E":          transform_echanges_externes,
    "Montée en charge":    transform_montee_en_charge,
    "Collaborateurs Clés": transform_collaborateurs,
    "Autres Eqt IT":       transform_autres_eqt,
    "Doc critiques":       transform_doc_critiques,
}


# ─────────────────────────────────────────────────────────────────────────────
# LOAD — XLSX WRITER
# ─────────────────────────────────────────────────────────────────────────────

HEADER_ROW = 5   # Row index (1-based) where column headers live in each sheet
DATA_START_ROW = 6  # First data row (1-based)

# Columns B, C, D hold Division / Unité / Département — used for row matching.
# We scan all three and pick the best fuzzy match.
DEPT_COLUMNS = [2, 3, 4]  # 1-based column indices (B=2, C=3, D=4)


def _find_department_row(ws, dept_name: str) -> Optional[int]:
    """
    Find which row in the sheet corresponds to dept_name.
    Uses rapidfuzz for fuzzy matching because the synthesis template has typos
    (e.g. "ingenieurue" instead of "ingénierie") and accent variations.
    Returns the 1-based row index, or None if not found.
    """
    candidates: dict[int, str] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        for col_idx in DEPT_COLUMNS:
            cell = ws.cell(row=row[0].row, column=col_idx)
            if cell.value and str(cell.value).strip() not in ("-", ""):
                candidates[row[0].row] = str(cell.value).strip()

    if not candidates:
        return None

    # rapidfuzz: token_sort_ratio handles word reordering, partial_ratio handles substrings
    best_row, best_score, _ = process.extractOne(
        dept_name, candidates,
        scorer=fuzz.token_sort_ratio,
    )
    if best_score < 60:
        return None
    # Return the row number (key in candidates)
    for row_num, name in candidates.items():
        if name == best_row:
            return row_num
    return None


def _get_column_map(ws) -> dict[str, int]:
    """Read header row → return {column_name: column_index (1-based)}."""
    col_map = {}
    for cell in ws[HEADER_ROW]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column
    return col_map


def _copy_row_style(ws, source_row: int, target_row: int):
    """Copy cell styles from source_row to target_row (Division/Unité/Département cols)."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        if src.has_style:
            tgt._style = copy.copy(src._style)


def load(synthesis_path: str | Path, fiche: BIAFiche, dry_run: bool = False,
         output_path: str | Path = None, verbose: bool = True):
    """
    LOAD phase: for each sheet in TRANSFORM_MAP, find the department row,
    insert extra rows if the fiche has multiple items (multiple activities,
    multiple apps, etc.), and fill in the data.

    Row insertion strategy:
      1. Find the template row for this department.
      2. If N items to insert > 1, insert (N-1) blank rows directly below,
         copying the row style so the sheet looks consistent.
      3. Fill all N rows with data starting from the template row.

    WHY insert instead of overwrite:
      The synthesis has one pre-existing row per department. When a department
      has 5 activities, we need 5 rows. We cannot merge into one cell because
      that breaks the tabular contract of the synthesis.
    """
    wb = load_workbook(str(synthesis_path))

    for sheet_name, transform_fn in TRANSFORM_MAP.items():
        # Normalize sheet name — the xlsx may have trailing spaces
        ws = None
        for sn in wb.sheetnames:
            if sn.strip() == sheet_name.strip():
                ws = wb[sn]
                break
        if ws is None:
            if verbose: print(f"  [WARN] Sheet '{sheet_name}' not found, skipping.")
            continue

        data_rows = transform_fn(fiche)
        if not data_rows:
            if verbose: print(f"  [INFO] No data for sheet '{sheet_name}' from this fiche.")
            continue

        col_map = _get_column_map(ws)
        dept_row = _find_department_row(ws, fiche.entity_name)

        if dept_row is None:
            if verbose: print(f"  [WARN] Department '{fiche.entity_name}' not found in sheet '{sheet_name}'.")
            continue

        n = len(data_rows)

        if not dry_run:
            # Insert extra rows if needed (N-1 because row dept_row already exists)
            if n > 1:
                ws.insert_rows(dept_row + 1, n - 1)
                for i in range(1, n):
                    _copy_row_style(ws, dept_row, dept_row + i)
                    # Copy Division/Unité/Département identifiers
                    for col in DEPT_COLUMNS:
                        ws.cell(row=dept_row + i, column=col).value = \
                            ws.cell(row=dept_row, column=col).value

            # Write data
            for i, data_row in enumerate(data_rows):
                target_row = dept_row + i
                for col_name, value in data_row.items():
                    # Fuzzy match column headers (handles minor label variations)
                    best_col_name = _fuzzy_col(col_name, col_map)
                    if best_col_name:
                        col_idx = col_map[best_col_name]
                        ws.cell(row=target_row, column=col_idx).value = value

        if verbose: print(f"  OK [{sheet_name}] -> {n} row(s) written for '{fiche.entity_name}'")

    if not dry_run:
        out_path = Path(output_path) if output_path else Path(synthesis_path).with_stem(
            Path(synthesis_path).stem + "_filled"
        )
        wb.save(str(out_path))
        if verbose: print(f"\nSaved: {out_path}")
    else:
        if verbose: print("\n[DRY RUN] No file written.")


def _fuzzy_col(col_name: str, col_map: dict[str, int]) -> Optional[str]:
    """Find the best matching column header in col_map using fuzzy matching."""
    if col_name in col_map:
        return col_name
    if not col_map:
        return None
    result = process.extractOne(col_name, list(col_map.keys()), scorer=fuzz.token_sort_ratio)
    if result and result[1] >= 70:
        return result[0]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def process_fiche(docx_path: Path, synthesis_path: Path, dry_run: bool):
    print(f"\n{'='*60}")
    print(f"Processing: {docx_path.name}")
    fiche = extract(docx_path)
    print(f"  Entity  : {fiche.entity_name}")
    print(f"  Activities    : {len(fiche.activities)}")
    print(f"  Impact rows   : {len(fiche.impact_rows)}")
    print(f"  Exchanges     : {len(fiche.exchanges)}")
    print(f"  IT Apps       : {len(fiche.it_applications)}")
    print(f"  Key people    : {len(fiche.key_people)}")
    print(f"  Ramp-up rows  : {len(fiche.ramp_up)}")
    print(f"  Other eqt     : {len(fiche.other_equipment)}")
    print(f"  Critical docs : {len(fiche.critical_docs)}")
    load(synthesis_path, fiche, dry_run=dry_run)


def main():
    parser = argparse.ArgumentParser(description="BIA ETL: fiches DOCX → synthèse XLSX")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--fiche", type=Path, help="Single BIA fiche .docx")
    group.add_argument("--fiches-dir", type=Path, help="Folder containing all BIA fiche .docx files")
    parser.add_argument("--synthese", type=Path, required=True, help="Synthèse BIA .xlsx template")
    parser.add_argument("--dry-run", action="store_true", help="Extract and print only, do not write")
    args = parser.parse_args()

    if not args.synthese.exists():
        sys.exit(f"ERROR: Synthèse file not found: {args.synthese}")

    if args.fiche:
        if not args.fiche.exists():
            sys.exit(f"ERROR: Fiche not found: {args.fiche}")
        process_fiche(args.fiche, args.synthese, args.dry_run)
    else:
        docx_files = sorted(args.fiches_dir.glob("*.docx"))
        if not docx_files:
            sys.exit(f"ERROR: No .docx files found in {args.fiches_dir}")
        print(f"Found {len(docx_files)} fiche(s) in {args.fiches_dir}")
        for f in docx_files:
            process_fiche(f, args.synthese, args.dry_run)


if __name__ == "__main__":
    main()
